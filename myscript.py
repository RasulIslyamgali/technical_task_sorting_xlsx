from openpyxl import load_workbook
import datetime
from openpyxl import Workbook
from pathlib import Path

wb = Workbook()
ws = wb.active
ws.title = 'Информация по операциям'

mains = {'mains': ("Компания", "Дата", "Тип операции", "Сумма")}
ws.append(mains['mains'])

# загружаем файлы
first_file = load_workbook(filename='first_table.xlsx')
second_file = load_workbook(filename='second_table.xlsx')


# выбираем таблицы
first_file_in = first_file['Информация по операциям']
second_file_in = second_file['Информация по операциям']


# Даты
choice_first = first_file_in['B']
choice_second = second_file_in['B']


# Компания
choice_first_company = first_file_in['A']
choice_second_company = second_file_in['A']


# Тип операции
text = 'Тип операции'
choice_oper_type_1 = first_file_in['C']
choice_oper_type_2 = second_file_in['C']


# price
choice_price_1 = first_file_in['D']
choice_price_2 = second_file_in['D']


# создаем словарь с данными, добавляем в список, сортируем по дате
list_first = []
# начало итерации с 1, чтобы не взять строк Компания, Дата, Тип операции , Сумма
for i in range(1, len(choice_first_company)):
    d_1 = dict(company_name=choice_first_company[i].value, date_time=choice_first[i].value,
               oper_type=choice_oper_type_1[i].value, price_=choice_price_1[i].value)
    d_2 = dict(company_name=choice_second_company[i].value, date_time=choice_second[i].value,
               oper_type=choice_oper_type_2[i].value, price_=choice_price_2[i].value)
    list_first.append(d_1)
    list_first.append(d_2)
sorted_list = sorted(
    list_first,
    key=lambda x: datetime.datetime.strptime(x['date_time'], '%Y-%m-%d %H:%M:%S'), reverse=False
)
# print(sorted_list)

for i in range(1, len(sorted_list) + 1):
    ws['A' + str(i+1)] = sorted_list[i - 1]['company_name']
    ws['B' + str(i+1)] = sorted_list[i - 1]['date_time']
    ws['C' + str(i+1)] = sorted_list[i -1]['oper_type']
    ws['D' + str(i+1)] = sorted_list[i - 1]['price_']

my_sorted_table = Path(r".\my_sorted_table.xlsx")

# будет 19999 полей, т.к. таблицы обьединены и 1 заглавочная строка не нужна
wb.save(my_sorted_table)
print('Done. Files was sorted')

